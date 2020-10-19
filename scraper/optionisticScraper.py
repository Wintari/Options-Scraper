from gevent import monkey
monkey.patch_all()

import sqlite3
import time
from lxml import html
import argparse
import random
import sys, os
import openpyxl
import re
import datetime
import requests
import gevent

def isNum(string):
    try:
        float(string)
        return True
    except ValueError:
        return False

def createKpi(option, prevOption, rootData, prevRootData, greeks, oldGreeks, curDate):
    date = curDate
    call = option[0]
    root = option[10]
    strike = option[12]
    put = option[14]

    cBid = option[1]
    cAsk = option[2]

    cPrice = option[3]
    try:
        dCPrice = float(option[3]) - prevOption[4]
    except:
        dCPrice = None
    
    try:
        dCPricePercent = round(dCPrice / prevOption[4] * 100, 2)
    except:
        dCPricePercent = None

    cTPrice = option[4]
    try:
        dCTPrice = float(option[4]) - prevOption[5]
    except:
        dCTPrice = None

    try:
        dCTPricePercent = round(dCTPrice / prevOption[5] * 100, 2)
    except:
        dCTPricePercent = None

    cVol = option[6]
    try:
        dCVol = float(option[6]) - prevOption[7]
    except:
        dCVol = None

    try:
        dCVolPercent = round(dCVol / prevOption[7] * 100, 2)
    except:
        dCVolPercent = None

    coi = option[7]
    try:
        dCoi = float(option[7]) - prevOption[8]
    except:
        dCoi = None

    try:
        dCoiPercent = round(dCoi / prevOption[8] * 100, 2)
    except:
        dCoiPercent = None

    cvolDivCoi = option[8]

    cns = option[9]
    cDelta = greeks[0]

    try:
        cMoneyness = None
        if(cDelta <= 0.98 and cDelta > 0.875):
            cMoneyness = 2
        elif(cDelta <= 0.875 and cDelta > 0.625):
            cMoneyness = 1
        elif(cDelta <= 0.625 and cDelta > 0.375):
            cMoneyness = 0
        elif(cDelta <= 0.375 and cDelta > 0.125):
            cMoneyness = -1
        elif(cDelta <= 0.125 and cDelta > 0.02):
            cMoneyness = -2
    except:
        cMoneyness = None 

    try:
        cMaturity = None
        callDate = datetime.datetime.strptime(call, "%B %d, %Y")
        curDate = datetime.datetime.now()
        maturity = (callDate - curDate).days

        if(maturity <= 7):
            cMaturity = 0
        elif(maturity <= 30):
            cMaturity = 1
        elif(maturity <= 60):
            cMaturity = 2
        elif(maturity > 60):
            cMaturity = 3 

    except:
        cMaturity = None

    CIVol = greeks[2]
    try:
        dCIVol = greeks[2] - oldGreeks[2]
    except:
        dCIVol = None

    root = rootData[0]
    sLast = rootData[2]
    sChange = rootData[3]
    precentSChange = rootData[4]
    strike = option[12]
    strikeDivSLast = option[13]

    impVol = rootData[5]
    try:
        dImpVol = float(rootData[5]) - prevRootData[6]
    except:
        dImpVol = None

    hisVol = rootData[6]
    try:
        dHisVol = float(rootData[6]) - prevRootData[7]
    except:
        dHisVol = None

    pBid = option[15]
    pAsk = option[16]

    pPrice = option[17]
    try:
        dPPrice = float(option[17]) - prevOption[18]
    except:
        dPPrice = None

    try:
        dPPricePercent = round(dPPrice / prevOption[18] * 100, 2)
    except:
        dPPricePercent = None

    pTPrice = option[18]
    try:
        dPTPrice = float(option[18]) - prevOption[19]
    except:
        dPTPrice = None

    try:
        dPTPricePercent = round(dPTPrice / prevOption[19] * 100, 2)
    except:
        dPTPricePercent = None

    pVol = option[20]
    try:
        dPVol = float(option[20]) - prevOption[21]
    except:
        dPVol = None

    try:
        dPVolPercent = round(dPVol / prevOption[21] * 100, 2)
    except:
        dPVolPercent = None

    poi = option[21]
    try:
        dPoi = float(option[21]) - prevOption[22]
    except:
        dPoi = None

    try:
        dPoiPercent = round(dPoi / prevOption[22] * 100, 2)
    except:
        dPoiPercent = None

    pvolDivPoi = option[22]

    pns = option[23]
    pDelta = greeks[1]

    try:
        pMoneyness = None
        if(pDelta > -0.98 and pDelta <= -0.875):
            pMoneyness = 2
        elif(pDelta > -0.875 and pDelta <= -0.625):
            pMoneyness = 1
        elif(pDelta > -0.625 and pDelta <= -0.375):
            pMoneyness = 0
        elif(pDelta > -0.375 and pDelta <= -0.125):
            pMoneyness = -1
        elif(pDelta > -0.125 and pDelta <= -0.02):
            pMoneyness = -2
    except:
        pMoneyness = None 

    try:
        pMaturity = None
        callDate = datetime.datetime.strptime(put, "%B %d, %Y")
        curDate = datetime.datetime.now()
        maturity = (callDate - curDate).days

        if(maturity <= 7):
            pMaturity = 0
        elif(maturity <= 30):
            pMaturity = 1
        elif(maturity <= 60):
            pMaturity = 2
        elif(maturity > 60):
            pMaturity = 3 

    except:
        pMaturity = None

    PIVol = greeks[3]
    try:
        dPIVol = greeks[3] - oldGreeks[3]
    except:
        dPIVol = None

    pvolDivCvol = option[24]
    poiDivCoi = option[25]
    poiMinCoi = option[26]

    try:
        dCIVolMindPIVol = dCIVol - dPIVol
    except:
        dCIVolMindPIVol = None

    return [date, call, cBid, cAsk, cPrice, dCPrice,dCPricePercent, cTPrice, dCTPrice, dCTPricePercent,
            cVol, dCVol, dCVolPercent, coi, dCoi, dCoiPercent, cvolDivCoi, cns, cDelta, cMoneyness, cMaturity, 
            CIVol, dCIVol, root, sLast, sChange, precentSChange, strike, strikeDivSLast, impVol, dImpVol,
            hisVol, dHisVol, put, pBid, pAsk, pPrice, dPPrice, dPPricePercent, pTPrice, dPTPrice, dPTPricePercent,
            pVol, dPVol, dPVolPercent, poi, dPoi, dPoiPercent, pvolDivPoi, pns, pDelta, pMoneyness, 
            pMaturity, PIVol, dPIVol, pvolDivCvol, poiDivCoi, poiMinCoi, dCIVolMindPIVol]

class DB:
    connection = None
    cursor = None
    date = None
    alreadyGettedRoots = set()

    def connect(self, file):
        DB.connection = sqlite3.connect(file)
        DB.cursor = DB.connection.cursor()

        tree = getPage("date")

        date = tree.find('''.//*[@id="mainbody"]/div[1]/form/h1/select/option[1]''').text.split('-')

        DB.date = date[1] + '.' + date[0] + '.' + datetime.datetime.now().strftime("%Y")
        DB.date = datetime.datetime.strptime(DB.date, "%d.%m.%Y")

    def createTables(self):
        if(DB.cursor):
            try:
                DB.cursor.execute(
                '''CREATE TABLE "Calls.Greeks" (
                    "Date"	TEXT,
                    "Calls"	TEXT,
                    "Root"	TEXT,
                    "Strike"	REAL,
                    "CIVol"	REAL,
                    "CDelta"	REAL,
                    "CTheta"	REAL,
                    "CVega"	REAL,
                    "CGamma"	REAL,
                    "CRho"	REAL,
                    PRIMARY KEY("Date","Calls","Root","Strike")
                )'''
                )
            except sqlite3.DatabaseError as err:       
                print("Error: ", err)

            try:
                DB.cursor.execute(
                '''CREATE TABLE "Calls.Main" (
                    "Date"	TEXT,
                    "Calls"	TEXT,
                    "Root"	TEXT,
                    "Strike"	REAL,
                    "CBid"	REAL,
                    "CAsk"	REAL,
                    "CPrice"	REAL,
                    "CTPrice"	REAL,
                    "CVolume"	INTEGER,
                    "COI"	INTEGER,
                    "CNS"	TEXT,
                    PRIMARY KEY("Date","Calls","Root","Strike")
                )'''
                )
            except sqlite3.DatabaseError as err:       
                print("Error: ", err)
            
            try:
                DB.cursor.execute(
                '''CREATE TABLE "Puts.Greeks" (
                    "Date"	TEXT,
                    "Puts"	TEXT,
                    "Root"	TEXT,
                    "Strike"	REAL,
                    "PIVol"	REAL,
                    "PDelta"	REAL,
                    "PTheta"	REAL,
                    "PVega"	REAL,
                    "PGamma"	REAL,
                    "PRho"	REAL,
                    PRIMARY KEY("Date","Puts","Root","Strike")
                )'''
                )
            except sqlite3.DatabaseError as err:       
                print("Error: ", err)

            try:
                DB.cursor.execute(
                '''CREATE TABLE "Puts.Main" (
                    "Date"	TEXT,
                    "Puts"	TEXT,
                    "Root"	TEXT,
                    "Strike"	REAL,
                    "PBid"	REAL,
                    "PAsk"	REAL,
                    "PPrice"	REAL,
                    "PTPrice"	REAL,
                    "PVolume"	INTEGER,
                    "POI"	INTEGER,
                    "PNS"	TEXT,
                    CONSTRAINT "Option" PRIMARY KEY("Date","Puts","Root","Strike")
                )'''
                )
            except sqlite3.DatabaseError as err:       
                print("Error: ", err)
            
            try:
                DB.cursor.execute(
                '''CREATE TABLE "Roots" (
                    "Date"	TEXT,
                    "Root"	TEXT,
                    "Name"	TEXT,
                    "SLast"	REAL,
                    "SChange"	REAL,
                    "% SChange"	REAL,
                    "Implied Vol"	REAL,
                    "Historical Vol"	REAL,
                    "Imp.Vol-His.Vol"   REAL,
                    "Imp.Vol/His.Vol"   REAL,
                    PRIMARY KEY("Date","Root","Name")
                )'''
                )
            except sqlite3.DatabaseError as err:       
                print("Error: ", err)
            
            try:
                DB.cursor.execute(
                '''CREATE TABLE "Options.Greeks" (
                    "Date"	TEXT,
                    "Calls"	TEXT,
                    "CIVol"	REAL,
                    "CDelta"	REAL,
                    "CTheta"	REAL,
                    "CVega"	REAL,
                    "CGamma"	REAL,
                    "CRho"	REAL,
                    "Root"	TEXT,
                    "SLast"	REAL,
                    "Strike"	REAL,
                    "Strike/SLast"	REAL,
                    "Puts"	TEXT,
                    "PIVol"	REAL,
                    "PDelta"	REAL,
                    "PTheta"	REAL,
                    "PVega"	REAL,
                    "PGamma"	REAL,
                    "PRho"	REAL,
                    "CIVol-PIVol"	REAL,
                    "CIVol/PIVol"	REAL,
                    PRIMARY KEY("Date","Calls","Root","Strike", "Puts")
                )'''
                )
            except sqlite3.DatabaseError as err:       
                print("Error: ", err)

            try:
                DB.cursor.execute(
                '''CREATE TABLE "Options.Main" (
                    "Date"	TEXT,
                    "Calls"	TEXT,
                    "CBid"	REAL,
                    "CAsk"	REAL,
                    "CPrice"	REAL,
                    "CTPrice"	REAL,
                    "CP-CTP"    REAL,
                    "CVolume"	INTEGER,
                    "COI"	INTEGER,
                    "CVol/COI"    REAL,
                    "CNS"	TEXT,
                    "Root"	TEXT,
                    "SLast"	REAL,
                    "Strike"	REAL,
                    "Strike/SLast"	REAL,
                    "Puts"	TEXT,
                    "PBid"	REAL,
                    "PAsk"	REAL,
                    "PPrice"	REAL,
                    "PTPrice"	REAL,
                    "PP-PTP"    REAL,
                    "PVolume"	INTEGER,
                    "POI"	INTEGER,
                    "PVol/POI"    REAL,
                    "PNS"	TEXT,
                    "PVol/СVol"    INTEGER,
                    "POI/СOI"    REAL,
                    "POI-СOI"    REAL,
                    PRIMARY KEY("Date","Calls","Root","Strike", "Puts")
                )'''
                )
            except sqlite3.DatabaseError as err:       
                print("Error: ", err)

            try:
                DB.cursor.execute(
                '''CREATE TABLE "Options.KPI" (
                    "Date"	TEXT,
                    "Calls"	TEXT,
                    "CBid"	REAL,
                    "CAsk"	REAL,
                    "CPrice"    REAL,
                    "dCPrice"    REAL,
                    "dCPrice%"    REAL,
                    "CTPrice"    REAL,
                    "dCTPrice"    REAL,
                    "dCTPrice%"    REAL,
                    "CVolume"    REAL,
                    "dCVolume"    REAL,
                    "dCVolume%"    REAL,
                    "СOI"    REAL,
                    "dСOI"    REAL,
                    "dСOI%"    REAL,
                    "CVol/COI"    REAL,
                    "CNS"    TEXT,
                    "CDelta"    REAL,
                    "CMoneyness"    INTEGER,
                    "CMaturity"    INTEGER,
                    "CIvol"    REAL,
                    "dCIvol"    REAL,
                    "Root"	TEXT,
                    "SLast"	REAL,
                    "SChange"	REAL,
                    "%SChange"	REAL,
                    "Strike"	REAL,
                    "Strike/SLast"	REAL,
                    "Implied Vol"	REAL,
                    "dImplied Vol"	REAL,
                    "Historical Vol"	REAL,
                    "dHistorical Vol"	REAL,
                    "Puts"	TEXT,
                    "PBid"	REAL,
                    "PAsk"	REAL,
                    "PPrice"    REAL,
                    "dPPrice"    REAL,
                    "dPPrice%"    REAL,
                    "PTPrice"    REAL,
                    "dPTPrice"    REAL,
                    "dPTPrice%"    REAL,
                    "PVolume"    REAL,
                    "dPVolume"    REAL,
                    "dPVolume%"    REAL,
                    "POI"    REAL,
                    "dPOI"    REAL,
                    "dPOI%"    REAL,
                    "PVol/POI"    REAL,
                    "PNS"    TEXT,
                    "PDelta"    REAL,
                    "PMoneyness"    INTEGER,
                    "PMaturity"    INTEGER,
                    "PIvol"    REAL,
                    "dPIvol"    REAL,
                    "PVol/СVol"    INTEGER,
                    "POI/СOI"    REAL,
                    "POI-СOI"    REAL,
                    "(dCIvol-dPIvol)" REAL,

                    PRIMARY KEY("Date","Calls","Root","Strike","Puts")
                )'''
                )
            except sqlite3.DatabaseError as err:       
                print("Error: ", err)

    def createLookup(self):
        print('Creating lookup...')

        try:
            DB.cursor.execute(
            '''DROP TABLE "Lookup"'''
            )
        except sqlite3.DatabaseError as err:       
            print("Error: ", err)
        try:
            DB.cursor.execute(
            '''CREATE TABLE "Lookup"
                AS SELECT opt.Calls, opt.Root, rt.Name, opt.Strike, opt.Puts
                FROM "Options.Main" as opt
                JOIN "Roots" as rt
                ON opt.Date = rt.Date
                AND opt.Root = rt.Root
                WHERE opt.Date = (?)''', [DB.date]
            )
        except sqlite3.DatabaseError as err:       
            print("Error: ", err)

        print('Lookup created.')

    def addKPIs(self, options, rootData, date = None):
        if(not date):
            date = DB.date

        if(DB.cursor and options):
            curDate = date
            prevDate = None
            prevDate = curDate - datetime.timedelta(days=1)

            root = rootData[0]

            try:
                DB.cursor.execute('''SELECT *
                                        FROM "Roots"
                                        WHERE Date = "%s" 
                                        AND Root = "%s";''' % (prevDate, root))

            except sqlite3.DatabaseError as err:       
                print("Error: ", err)

            prevRootData = DB.cursor.fetchone()
            
            for option in options:
                call = option[0]
                strike = option[12]
                put = option[14]

                try:
                    DB.cursor.execute('''SELECT *
                                        FROM "Options.Main"
                                        WHERE Date = "%s" 
                                        AND Calls = "%s"
                                        AND Root = "%s"
                                        AND Strike = "%s"
                                        AND Puts = "%s";''' % (prevDate, call, root, strike, put))
                except sqlite3.DatabaseError as err:       
                    print("Error: ", err)

                prevOption = DB.cursor.fetchone()

                try:
                    DB.cursor.execute('''SELECT CDelta, PDelta, CIVol, PIVol
                                        FROM "Options.Greeks"
                                        WHERE Date = "%s" 
                                        AND Calls = "%s"
                                        AND Root = "%s"
                                        AND Strike = "%s"
                                        AND Puts = "%s";''' % (curDate, call, root, strike, put))
                except sqlite3.DatabaseError as err:       
                    print("Error: ", err)

                greeks = DB.cursor.fetchone()

                try:
                    DB.cursor.execute('''SELECT CDelta, PDelta, CIVol, PIVol
                                        FROM "Options.Greeks"
                                        WHERE Date = "%s" 
                                        AND Calls = "%s"
                                        AND Root = "%s"
                                        AND Strike = "%s"
                                        AND Puts = "%s";''' % (prevDate, call, root, strike, put))
                except sqlite3.DatabaseError as err:       
                    print("Error: ", err)

                oldGreeks = DB.cursor.fetchone()

                kpi = createKpi(option, prevOption, rootData, prevRootData, greeks, oldGreeks, curDate)

                try:
                    DB.cursor.execute('''insert into "Options.KPI" values (?, ?, ?, ?, ?, ?, ?, ?,
                                        ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                                        ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                                        ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);''', kpi)
                except sqlite3.DatabaseError as err:       
                    print("Error: ", err)

    def addOptions(self, options, date = None):
        if(not date):
            date = DB.date

        if(DB.cursor and options):
            for option in options:
                call = [date, option[0], option[10], option[12], option[1], option[2], option[3], option[4], option[6], option[7], option[9]]
                put = [date, option[14], option[10], option[12], option[15], option[16], option[17], option[18], option[20], option[21], option[23]]
                try:
                    DB.cursor.execute('''insert into "Calls.Main" values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);''', call)
                except sqlite3.DatabaseError as err:       
                    print("Error: ", err)
                try:
                    DB.cursor.execute('''insert into "Puts.Main" values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);''', put)
                except sqlite3.DatabaseError as err:       
                    print("Error: ", err)
                try:
                    fullOption = [date]
                    fullOption.extend(option)
                    DB.cursor.execute('''insert into "Options.Main" values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);''', fullOption)
                except sqlite3.DatabaseError as err:       
                    print("Error: ", err)

    def addGreeks(self, greeks, date = None):
        if(not date):
            date = DB.date

        if(DB.cursor and greeks):
            for greek in greeks:
                call = [date, greek[0], greek[7], greek[9], greek[1], greek[2], greek[3], greek[4], greek[5], greek[6]]
                put = [date, greek[11], greek[7], greek[9], greek[12], greek[13], greek[14], greek[15], greek[16], greek[17]]
                try:
                    DB.cursor.execute('''insert into "Calls.Greeks" values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);''', call)
                except sqlite3.DatabaseError as err:       
                    print("Error: ", err)
                try:
                    DB.cursor.execute('''insert into "Puts.Greeks" values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);''', put)
                except sqlite3.DatabaseError as err:       
                    print("Error: ", err)
                try:
                    fullGreek = [date]
                    fullGreek.extend(greek)
                    DB.cursor.execute('''insert into "Options.Greeks" values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);''', fullGreek)
                except sqlite3.DatabaseError as err:       
                    print("Error: ", err)

    
    def addRoot(self, rootData, date = None):
        if(not date):
            date = DB.date

        if(DB.cursor and rootData):
            root = [date]
            root.extend(rootData)
            try:
                DB.cursor.execute('''insert into "Roots" values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);''', root)
            except sqlite3.DatabaseError as err:       
                print("Error: ", err)

    def fillOptions(self, options):
        print('Filling option table...')
        if(DB.cursor):
            dates = [DB.date + datetime.timedelta(days=-9 + i) for i in range(9)]
            for date in dates:
                datedOption = []
                for option in options:
                    last = [option[0], 0, 0, 0, 0, 0,
                        0, 0, None, None, option[10], 0, option[12], None, option[14], 0, 0,
                        0, 0, 0, 0, 0, None, None, None, None, 0]
                    try:
                        DB.cursor.execute('''SELECT *
                                    FROM "Options.Main"
                                    WHERE Date = "%s" 
                                    AND Calls = "%s"
                                    AND Root = "%s"
                                    AND Strike = "%s"
                                    AND Puts = "%s";''' % (date, option[0], option[10], option[12], option[14]))
                    except sqlite3.DatabaseError as err:       
                        print("Error: ", err)

                    old = DB.cursor.fetchone()
                    if(old):
                        last = old[1:]
                    else:
                        datedOption.append(last)

                self.addOptions(datedOption, date)

        print('Option table filled...')
        

    def fillGreeks(self, greeks):
        print('Filling greeks table...')
        if(DB.cursor):
            dates = [DB.date + datetime.timedelta(days=-9 + i) for i in range(9)]

            for date in dates:
                datedGreeks = []
                for greek in greeks:
                    last = [greek[0], 0, 0, 0, 0, 0,
                        0, greek[7], 0, greek[9], None, greek[11], 0, 0,
                        0, 0, 0, 0, None, None]
                    try:
                        DB.cursor.execute('''SELECT *
                                    FROM "Options.Greeks"
                                    WHERE Date = "%s" 
                                    AND Calls = "%s"
                                    AND Root = "%s"
                                    AND Strike = "%s"
                                    AND Puts = "%s";''' % (date, greek[0], greek[7], greek[9], greek[11]))
                    except sqlite3.DatabaseError as err:       
                        print("Error: ", err)

                    old = DB.cursor.fetchone() 
                    if(old):
                        last = old[1:]
                    else:
                        datedGreeks.append(last)
                self.addGreeks(datedGreeks, date)
        print('Greeks table filled...')

    def fillRoot(self, root):
        print('Filling root table...')
        if(DB.cursor):
            dates = [DB.date + datetime.timedelta(days=-9 + i) for i in range(9)]
            last = [root[0], root[1], 0, 0, None, 0, 0, 0, None]
            for date in dates:
                try:
                    DB.cursor.execute('''SELECT *
                                FROM "Roots"
                                WHERE Date = "%s" 
                                AND Root = "%s";''' % (date, root[0]))
                except sqlite3.DatabaseError as err:       
                    print("Error: ", err)

                old = DB.cursor.fetchone() 
                if(old):
                    last = old[1:]
                else:
                    self.addRoot(last, date)
        print('Root table filled...')

    def fillKPIs(self, options, root):
        print('Filling KPIs table...')
        if(DB.cursor):
            dates = [DB.date + datetime.timedelta(days=-9 + i) for i in range(9)]

            for date in dates:
                datedOptions = []
                lastRoot = [root[0], root[1], 0, 0, None, 0, 0, 0, None]
                try:
                    DB.cursor.execute('''SELECT *
                                FROM "Roots"
                                WHERE Date = "%s" 
                                AND Root = "%s";''' % (date, root[0]))
                except sqlite3.DatabaseError as err:       
                    print("Error: ", err)

                oldRoot = DB.cursor.fetchone() 
                if(oldRoot):
                        lastRoot = oldRoot[1:]

                for option in options:
                    lastOption = [option[0], 0, 0, 0, 0, 0,
                        0, 0, None, None, option[10], 0, option[12], None, option[14], 0, 0,
                        0, 0, 0, 0, 0, None, None, None, None, 0]
                    try:
                        DB.cursor.execute('''SELECT *
                                    FROM "Options.Main"
                                    WHERE Date = "%s" 
                                    AND Calls = "%s"
                                    AND Root = "%s"
                                    AND Strike = "%s"
                                    AND Puts = "%s";''' % (date, option[0], option[10], option[12], option[14]))
                    except sqlite3.DatabaseError as err:       
                        print("Error: ", err)

                        oldOption = DB.cursor.fetchone()

                    try:
                        DB.cursor.execute('''SELECT *
                                    FROM "Options.KPI"
                                    WHERE Date = "%s" 
                                    AND Calls = "%s"
                                    AND Root = "%s"
                                    AND Strike = "%s"
                                    AND Puts = "%s";''' % (date, option[0], option[10], option[12], option[14]))
                    except sqlite3.DatabaseError as err:       
                        print("Error: ", err)

                    oldKPI = DB.cursor.fetchone()

                    if(oldKPI):
                        if(oldOption):
                            lastOption = oldOption[1:]
                    else:
                        datedOptions.append(lastOption)

                self.addKPIs(datedOptions, lastRoot, date)

        print('KPIs table filled...')

    def isAlreadyGettedToday(self, root):
        if(len(DB.alreadyGettedRoots) == 0):
            try:
                DB.cursor.execute('''select DISTINCT Root from "Roots" where Date = (?)''', [DB.date])
                roots = DB.cursor.fetchall()
            except sqlite3.DatabaseError as err:
                print("Error: ", err)
            else:
                DB.alreadyGettedRoots.update((None,))
                for elem in roots:
                    DB.alreadyGettedRoots.update(elem)
                
        return (root in DB.alreadyGettedRoots or root.upper() in DB.alreadyGettedRoots or root.lower() in DB.alreadyGettedRoots)

class Proxies:
    proxies = []
    def getProxy(self):
        if(Proxies.proxies):
            return Proxies.proxies[random.randint(0, len(Proxies.proxies) - 1)]
        else:
            print('No valid proxies')
            return None

    def loadProxies(self, path):
        with open(path, 'r') as file:
            threads = []
            for line in file:
                threads.append(gevent.spawn(validateProxy, line.replace('\n', '')))
    
            gevent.joinall(threads)
            for thread in threads:
                if(thread.value != None):
                    Proxies.proxies.append(thread.value)

        print(str(len(Proxies.proxies)) + ' valid proxies at start')

    def invalidProxy(self, proxy):
        try:
            Proxies.proxies.remove(proxy)
        except:
            pass
        else:
            print(proxy + ' removed. Remaining proxies: ' + str(len(Proxies.proxies)))

def validateProxy(proxy):
    print('Validating proxy: ' + proxy)
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.135 Safari/537.36'}
        response = requests.get('https://www.optionistics.com/', headers = headers, proxies={'http' : 'http://' + proxy, 'https' : 'http://' + proxy}, timeout=5)
        tree = html.document_fromstring(response.content)
    except:
        print(proxy + ' - bad proxy ')
        return None
    else:
        siteName = tree.find('''.//body/div/table/tr[1]/td/table/tr/td[1]/table/tr[1]/td[1]/a/img''')
        if(siteName == None):
            print(proxy + ' - banned')
            return None
    print(proxy + ' - valid')
    return proxy

class ParsingStatus:
    rootsCount = 0
    rootsParsed = 0
    rootsInvalid = []
    def setRootCount(self, count):
        ParsingStatus.rootsCount = count

    def parsingStarted(self, root):
        print('Root started: ' + root)

    def parsingFinished(self, root):
        ParsingStatus.rootsParsed += 1
        print(f'\rRoot finished: ' + root + '. Roots parsed: ' + str(ParsingStatus.rootsParsed) + '/' + str(ParsingStatus.rootsCount))

    def invalidRoot(self, root):
        ParsingStatus.rootsInvalid.append(root)
        print('Invalid root: ' + root)

    def failedRoots(self):
        return  ParsingStatus.rootsInvalid

def getPage(root):
    proxies = Proxies()

    timeout = 5

    while(True):
        try:
            proxy = proxies.getProxy()
            print(' Getting root: ' + root + ' as ' + proxy)
            if(proxy):
                headers = {
                    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.121 Safari/537.36'}
                requestData = {
                    'symbol': root,
                    'range': '-1',
                    'putcall': '-1',
                    'expiry': '-1',
                    'strike': '',
                    'nonstd': '-1',
                    'greeks': '1',
                    'from': '0',
                    'sub': '0',
                    'lock': '0',
                    'v': '1',
                    'parm': ''}

                response = requests.post('''https://www.optionistics.com/f/option_chains''', requestData, headers=headers, proxies={'http' : 'http://' + proxy, 'https' : 'http://' + proxy}, timeout = timeout)

                tree = html.document_fromstring(response.content)
        except requests.exceptions.Timeout:
            print(' Timeout %s sec. Retrying: ' % timeout + root)
            timeout = timeout * 2
        except:
            pass
        else:
            siteName = tree.find('''.//body/div/table/tr[1]/td/table/tr/td[1]/table/tr[1]/td[1]/a/img''')
            if(siteName == None):
                print(' Invalid page. Retrying.: ' + root)
            else:
                print(' Root getted: ' + root)
                break

    return tree


def parsePage(tree, root, last):
    dates = tree.findall('''.//*[@id="mainbody"]/div[2]/table[2]/tr''')
    options = []
    greeks = []

    for date in dates:
        i = 0
        call = ''
        put = ''
        strike = 0
        for row in date.findall('''./td/table/tr'''):
            textRow = []

            for elem in row.findall('''./td'''):
                text = elem.text
                if(text):
                    text = text.replace(u"\xa0", u"")

                textRow.append(text)

            if(i == 0):
                call = textRow[0].split(' - ')[0][1:]
                put = textRow[2].split(' - ')[0][1:]
            elif(i > 2):
                if(i % 2 == 1):
                    strike = textRow[8]

                    try:
                        cpMinCtp = str(round(float(textRow[3]) - float(textRow[4]), 2))
                    except:
                        cpMinCtp = None

                    try:
                        ppMinPtp = str(round(float(textRow[12]) - float(textRow[13]), 2))
                    except:
                        ppMinPtp = None

                    try:
                        pvolDivCvol = str(round(float(textRow[14]) / float(textRow[5]), 2))
                    except:
                        pvolDivCvol = None

                    try:
                        poiDivCoi = str(round(float(textRow[15]) / float(textRow[6]), 2))
                    except:
                        poiDivCoi = None

                    try:
                        poiMinCoi = str(round(float(textRow[15]) - float(textRow[6]), 2))
                    except:
                        poiMinCoi = None

                    try:
                        strikeDivSlast = str(round(float(strike) / float(last), 2))
                    except:
                        strikeDivSlast = None

                    try:
                        cvolDivCoi = str(round(float(textRow[5]) / float(textRow[6]), 2))
                    except:
                        cvolDivCoi = None

                    try:
                        pvolDivPoi = str(round(float(textRow[14]) / float(textRow[15]), 2))
                    except:
                        pvolDivPoi = None

                    option = [call, textRow[1], textRow[2], textRow[3], textRow[4], cpMinCtp,
                    textRow[5], textRow[6], cvolDivCoi, textRow[7], root, last, strike, strikeDivSlast, put, textRow[10], textRow[11],
                    textRow[12], textRow[13], ppMinPtp, textRow[14], textRow[15], pvolDivPoi, textRow[16], pvolDivCvol, poiDivCoi, poiMinCoi]

                    options.append(option)
                else:
                    
                    try:
                        civolMinPivol = str(round(float(textRow[1]) - float(textRow[9]), 2))
                    except:
                        civolMinPivol = None

                    try:
                        civolDivPivol = str(round(float(textRow[1]) / float(textRow[9]), 2))
                    except:
                        civolDivPivol = None

                    try:
                        strikeDivSlast = str(round(float(strike) / float(last), 2))
                    except:
                        strikeDivSlast = None

                    greek = [call]
                    isPartSkipped = False

                    if(textRow[1] == 'Greeks could not be computed'):
                        greek.extend([None, None, None, None, None, None])
                        isPartSkipped = True
                    else:
                        greek.extend([textRow[1], textRow[2], textRow[3], textRow[4],
                    textRow[5], textRow[6]])

                    greek.extend([root, last, strike, strikeDivSlast, put])

                    if(isPartSkipped):
                        if(textRow[3] == 'Greeks could not be computed'):
                            greek.extend([None, None, None, None, None, None])
                        else:
                            greek.extend([textRow[3], textRow[4], textRow[5], textRow[6],
                        textRow[7], textRow[8]])
                    else:
                        if(textRow[9] == 'Greeks could not be computed'):
                            greek.extend([None, None, None, None, None, None])
                        else:
                            greek.extend([textRow[9], textRow[10], textRow[11], textRow[12],
                        textRow[13], textRow[14]])

                    greek.extend([civolMinPivol, civolDivPivol])
                    
                    greeks.append(greek)

            i += 1

    return options, greeks

def parseRoot(root):
    root = root.upper().replace(' ', '.')
    status = ParsingStatus()
    db = DB()

    if(db.isAlreadyGettedToday(root)):
        print(' Root: ' + root + ' already getted.')
        status.parsingFinished(root=root)
        return

    tree = getPage(root)

    rootName = tree.find('''.//*[@id="mainbody"]/div[2]/table[1]/tr[3]/td[2]/b''')

    if(rootName != None):
        rootName = rootName.text.replace(u"\xa0", u"")
        impVol = tree.find('''.//*[@id="mainbody"]/div[2]/table[1]/tr[3]/td[4]/b''').text
        histVol = tree.find('''.//*[@id="mainbody"]/div[2]/table[1]/tr[3]/td[5]/b''').text
        price = tree.find('''.//*[@id="mainbody"]/div[2]/table[1]/tr[3]/td[6]/b''').text
        try:
            change = tree.find('''.//*[@id="mainbody"]/div[2]/table[1]/tr[3]/td[7]/b/font''').text
        except:
            change = tree.find('''.//*[@id="mainbody"]/div[2]/table[1]/tr[3]/td[7]/b''').text

        try:
            percentChange = str(round(float(change) / float(price) * 100, 2))
        except:
            percentChange = None

        try:
            impMinHis = str(round(float(impVol) - float(histVol), 2))
        except:
            impMinHis = None
        
        try:
            impDivHis = str(round(float(impVol) / float(histVol), 2))
        except:
            impDivHis = None

        if(rootName != 'Unknown security'):
            options, greeks = parsePage(tree, root, price)
            rootData = [root, rootName, price, change, percentChange, impVol, histVol, impMinHis, impDivHis]

            print('Updating options table...')
            db.addOptions(options)
            db.fillOptions(options)
            print('Options table updated')
            print('Updating greeks table...')
            db.addGreeks(greeks)
            db.fillGreeks(greeks)
            print('Greeks table updated')
            print('Updating roots table...')
            db.addRoot(rootData)
            db.fillRoot(rootData)
            db.connection.commit()
            print('Roots table updated')
            print('Updating KPIs table...')
            db.addKPIs(options, rootData)
            db.fillKPIs(options, rootData)
            DB.connection.commit()
            print('KPIs table updated')

        else:
            status.invalidRoot(root)

    status.parsingFinished(root)


class Parser:
    queue = gevent.queue.Queue()

    def addRoots(self, roots):
        queue = []
        for root in roots:
            if(root):
                queue.append(root)
        self.queue = gevent.queue.Queue(items=queue)

        status = ParsingStatus()
        status.setRootCount(len(queue))

    def runThread(self):
        while(True):
            try:
                parseRoot(self.queue.get(block=False))
            except gevent.queue.Empty:
                return

    def run(self, threadsCount):
        threads = []

        for i in range(threadsCount):
            threads.append(gevent.spawn(self.runThread))
    
        gevent.joinall(threads)


def main():
    start_time = datetime.datetime.now()
    parser = argparse.ArgumentParser()
    parser.add_argument('-db')
    parser.add_argument('-proxies')
    parser.add_argument('-roots')
    args = parser.parse_args()

    proxies = Proxies()
    proxies.loadProxies(args.proxies)

    db = DB()
    db.connect(args.db)
    db.createTables()

    roots = []
    wb = openpyxl.load_workbook(args.roots)
    wb.active = 0
    sheet = wb.active
    for i in range(4, sheet.max_row + 1):
        roots.append(sheet.cell(row = i, column = 2).value)

    parser  = Parser()
    parser.addRoots(roots)
    parser.run(len(roots))

    print('All roots finished')

    db.createLookup()

    print(datetime.datetime.now() - start_time)
    
if __name__ == "__main__":
    main()

